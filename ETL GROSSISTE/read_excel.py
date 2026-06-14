from openpyxl import load_workbook
import sys

try:
    wb = load_workbook('D:/reporting-module/DATA BIBLOS/REPORTING DES VENTES ACTIVATION GROSSISTE BIBLOS INDUST 5.xlsx')
    print('Sheet names:', wb.sheetnames)
    ws = wb.active
    print('Active sheet:', ws.title)
    
    # Get first sheet
    first_sheet = wb[wb.sheetnames[0]]
    print('First sheet name:', first_sheet.title)
    
    # Read first 20 rows
    print('\nFirst 20 rows:')
    for i, row in enumerate(first_sheet.iter_rows(max_row=20, values_only=True)):
        print(f'Row {i}:', row)
        
except Exception as e:
    print(f'Error: {e}', file=sys.stderr)
    import traceback
    traceback.print_exc()
