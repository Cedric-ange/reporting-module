import sys
import traceback

try:
    print("Starting to read Excel file...", file=sys.stderr)
    
    # First try with openpyxl
    from openpyxl import load_workbook
    print("openpyxl imported successfully", file=sys.stderr)
    
    file_path = 'D:/reporting-module/DATA BIBLOS/REPORTING DES VENTES ACTIVATION GROSSISTE BIBLOS INDUST 5.xlsx'
    print(f"Attempting to load: {file_path}", file=sys.stderr)
    
    wb = load_workbook(file_path, data_only=True)
    print(f"Workbook loaded. Sheet names: {wb.sheetnames}", file=sys.stderr)
    
    # Get first sheet
    first_sheet = wb[wb.sheetnames[0]]
    print(f"First sheet: {first_sheet.title}", file=sys.stderr)
    
    # Read first 30 rows
    print("\n=== FIRST 30 ROWS ===")
    for i, row in enumerate(first_sheet.iter_rows(max_row=30, values_only=True)):
        print(f"Row {i}: {row}")
        
    print("\n=== SUCCESS ===")
    
except Exception as e:
    print(f"ERROR: {e}", file=sys.stderr)
    traceback.print_exc()
