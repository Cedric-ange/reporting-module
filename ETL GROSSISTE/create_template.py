import shutil
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import os

# Chemins des fichiers
source_file = r'C:\Users\angec\Downloads\SAN PEDRO - REPORTING COMMANDO DE REFE. DETAILLANTS LAIT & FLOC. AVOINE (1).xlsx'
dest_file = r'C:\Users\angec\reporting-module\backend\templates\template_original.xlsx'

# Copier le fichier original
shutil.copy(source_file, dest_file)

# Charger le workbook
wb = load_workbook(dest_file)

# Nettoyer les données mais préserver la structure, en-têtes, couleurs et formules
for sheet_name in wb.sheetnames:
    if sheet_name == 'Synthèse semaine':
        continue  # Garder la synthèse telle quelle
    
    ws = wb[sheet_name]
    
    # Identifier la ligne où commencent les données (après les en-têtes)
    data_start_row = None
    for row_idx, row in enumerate(ws.iter_rows()):
        if row_idx == 0:
            continue  # Première ligne vide
        
        # Chercher la première ligne de données (qui contient un numéro d'agent)
        first_cell = row[0].value
        if first_cell and isinstance(first_cell, (int, float)):
            data_start_row = row_idx + 1  # +1 car iter_rows est 0-indexed
            break
    
    if data_start_row:
        # Nettoyer les données à partir de cette ligne
        for row_idx in range(data_start_row, ws.max_row + 1):
            for col_idx in range(ws.max_column):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Nettoyer les valeurs mais préserver les formules et le formatage
                if cell.value and not cell.data_type == 'f':  # Si ce n'est pas une formule
                    # Pour les colonnes numériques, mettre à 0
                    if isinstance(cell.value, (int, float)) and cell.value != 0:
                        cell.value = 0
                    # Pour les colonnes de texte, laisser vide
                    elif isinstance(cell.value, str) and cell.value.strip():
                        if col_idx == 1:  # Colonne du nom d'agent
                            cell.value = ''  # Garder vide
                # Préserver les formules et le formatage

# Sauvegarder le template nettoyé
wb.save(dest_file)

print(f"Template créé: {dest_file}")
print(f"Feuilles: {wb.sheetnames}")
print("Structure, en-têtes, couleurs et formules préservées")
print("Données de vente/nettoyées, prêtes pour remplissage")