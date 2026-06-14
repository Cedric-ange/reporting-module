import sys
import traceback

try:
    from openpyxl import load_workbook
    
    file_path = 'D:/reporting-module/DATA BIBLOS/REPORTING DES VENTES ACTIVATION GROSSISTE BIBLOS INDUST 5.xlsx'
    wb = load_workbook(file_path, data_only=True)
    
    first_sheet = wb[wb.sheetnames[0]]
    print(f"Sheet name: {first_sheet.title}")
    
    # Read more rows to understand full structure
    print("\n=== FULL STRUCTURE (First 200 rows) ===")
    for i, row in enumerate(first_sheet.iter_rows(max_row=200, values_only=True)):
        # Print all rows but mark empty ones
        if any(cell is not None for cell in row):
            print(f"Row {i}: {row}")
        else:
            print(f"Row {i}: [EMPTY]")
        
except Exception as e:
    print(f"ERROR: {e}", file=sys.stderr)
    traceback.print_exc()
