import sys
import traceback

try:
    from openpyxl import load_workbook
    
    file_path = 'D:/reporting-module/DATA BIBLOS/REPORTING DES VENTES ACTIVATION GROSSISTE BIBLOS INDUST 5.xlsx'
    wb = load_workbook(file_path, data_only=True)
    
    first_sheet = wb[wb.sheetnames[0]]
    print(f"Sheet name: {first_sheet.title}")
    
    # Read first 100 rows to understand structure
    print("\n=== FIRST 100 ROWS ===")
    for i, row in enumerate(first_sheet.iter_rows(max_row=100, values_only=True)):
        # Filter out completely None rows for clarity
        if any(cell is not None for cell in row):
            print(f"Row {i}: {row}")
        
except Exception as e:
    print(f"ERROR: {e}", file=sys.stderr)
    traceback.print_exc()
